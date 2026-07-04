import telebot
import openpyxl
from datetime import datetime
import os
from fuzzywuzzy import fuzz
from flask import Flask, request
import json

# ============ ENVIRONMENT VARIABLES ============
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
WEBHOOK_URL = os.getenv("WEBHOOK_URL")

print("=" * 70)
print("🚀 BOT INITIALIZATION")
print("=" * 70)
print(f"TOKEN available: {bool(TOKEN)}")
if TOKEN:
    print(f"TOKEN: {TOKEN[:20]}...{TOKEN[-5:]}")
print(f"WEBHOOK_URL: {WEBHOOK_URL}")
print("=" * 70)

bot = telebot.TeleBot(TOKEN, parse_mode="Markdown")
app = Flask(__name__)
user_data = {}

def ismlarni_standartlash(ism):
    if not ism: return ""
    ism = str(ism).strip().lower()
    ism = ism.replace("`", "").replace("ʻ", "").replace("'", "").replace("'", "").replace("'", "")
    ism = ism.replace("о'", "o").replace("o'", "o").replace("o'", "o").replace("о'", "o")
    ism = ism.replace("g'", "g").replace("g'", "g").replace("г'", "g")
    ism = ism.replace("ch", "c").replace("sh", "s")
    ism = ism.replace("x", "h").replace("ya", "a").replace("yu", "u")
    ism = ism.replace("i", "e").replace("a", "e")
    return "".join(ism.split())

# ============ BOT MESSAGE HANDLERS ============
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    chat_id = message.chat.id
    print(f"✅ /START COMMAND RECEIVED from {chat_id}")
    user_data[chat_id] = {"holat": "sana_kutish"}
    
    response = ("Salom! Render bulutida 24/7 ishlovchi aqlli tizim faol. 🚀\n\n"
                "To'lovlarni **qaysi sanadan boshlab** hisoblayiy?\n"
                "Format: `27.06.2026` shaklida yozing:")
    bot.send_message(chat_id, response)
    print(f"✅ RESPONSE SENT to {chat_id}")

@bot.message_handler(func=lambda message: message.chat.id in user_data and user_data[message.chat.id].get("holat") == "sana_kutish")
def qabul_qilish_sanasi(message):
    chat_id = message.chat.id
    sana_matni = message.text.strip()
    print(f"📅 DATE RECEIVED: {sana_matni} from {chat_id}")
    try:
        cheklov_sanasi = datetime.strptime(sana_matni, "%d.%m.%Y")
        user_data[chat_id]["sana"] = cheklov_sanasi
        user_data[chat_id]["holat"] = "fayl_kutish"
        response = (f"✅ Sana tasdiqlandi: **{sana_matni}**.\n\n"
                    f"1. **Asosiy bazangizni** (.xlsx) yuboring.")
        bot.send_message(chat_id, response)
        print(f"✅ DATE CONFIRMED for {chat_id}")
    except ValueError:
        print(f"❌ INVALID DATE FORMAT: {sana_matni}")
        bot.send_message(chat_id, "❌ Noto'g'ri format. Nuqtalar bilan kiriting (Masalan: 27.06.2026):")

@bot.message_handler(content_types=['document'])
def handle_docs(message):
    chat_id = message.chat.id
    file_name = message.document.file_name if message.document else "unknown"
    print(f"📄 DOCUMENT RECEIVED: {file_name} from {chat_id}")
    
    if chat_id not in user_data or "sana" not in user_data[chat_id]:
        print(f"❌ NO DATE SET for {chat_id}")
        bot.send_message(chat_id, "❌ Iltimos, avval /start buyrug'ini bosing!")
        return

    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        baza_nomi = f"baza_{chat_id}.xlsx"
        deb_nomi = f"deb_{chat_id}.xlsx"
        natija_nomi = f"Tayyor_Yangilangan_{chat_id}.xlsx"

        if "baza_yuklandi" not in user_data[chat_id]:
            with open(baza_nomi, 'wb') as f:
                f.write(downloaded_file)
            user_data[chat_id]["baza_yuklandi"] = True
            user_data[chat_id]["baza_path"] = baza_nomi
            print(f"✅ BASE FILE SAVED: {baza_nomi}")
            bot.send_message(chat_id, "✅ Asosiy baza yuklandi.\n\n2. Endi bankdan kelgan **Debitorka** faylini yuboring.")
            return
        
        elif "deb_yuklandi" not in user_data[chat_id]:
            with open(deb_nomi, 'wb') as f:
                f.write(downloaded_file)
            user_data[chat_id]["deb_yuklandi"] = True
            print(f"✅ DEBTOR FILE SAVED: {deb_nomi}")
            
            baza_path = user_data[chat_id]["baza_path"]
            cheklov_sanasi = user_data[chat_id]["sana"]
            
            bot.send_message(chat_id, "🔄 Hisob-kitob va matnli hisobot tayyorlanmoqda. Kuting...")

            wb_baza_write = openpyxl.load_workbook(baza_path, data_only=False)
            wb_baza_read = openpyxl.load_workbook(baza_path, data_only=True)
            wb_deb = openpyxl.load_workbook(deb_nomi, data_only=True)

            varoq_nomi = 'KONTRAKTLAR' if 'KONTRAKTLAR' in wb_baza_write.sheetnames else wb_baza_write.sheetnames[0]
            sheet_write = wb_baza_write[varoq_nomi]
            sheet_read = wb_baza_read[varoq_nomi]
            sheet_deb = wb_deb['bank']

            baza_talabalari = []
            for row in range(25, sheet_read.max_row + 1):
                fio = sheet_read.cell(row=row, column=3).value
                if fio and str(fio).strip() and not str(fio).lower().startswith(('familiya', 'f.i.sh', 'итого', 'jami')):
                    baza_talabalari.append({
                        "row": row,
                        "original_name": str(fio).strip(),
                        "clean_name": ismlarni_standartlash(fio)
                    })

            yangilanish_tarixi = []
            topilmaganlar = []
            
            for row in range(2, sheet_deb.max_row + 1):
                sana_val = sheet_deb.cell(row=row, column=1).value
                if not sana_val: continue

                try:
                    if isinstance(sana_val, str):
                        if '.' in sana_val:
                            to_lov_sanasi = datetime.strptime(sana_val.strip(), "%d.%m.%y")
                        else:
                            to_lov_sanasi = datetime.strptime(sana_val.strip(), "%Y-%m-%d %H:%M:%S")
                    elif isinstance(sana_val, datetime):
                        to_lov_sanasi = sana_val
                    else:
                        continue
                except ValueError:
                    continue

                if to_lov_sanasi >= cheklov_sanasi:
                    summa_val = sheet_deb.cell(row=row, column=7).value
                    deb_fio = sheet_deb.cell(row=row, column=9).value

                    if not deb_fio or not summa_val: continue

                    yangi_summa = float(summa_val)
                    deb_fio_clean = ismlarni_standartlash(deb_fio)

                    eng_yaxshi_moslik = None
                    eng_yuqori_ball = 0

                    for talaba in baza_talabalari:
                        ball = fuzz.token_sort_ratio(deb_fio_clean, talaba["clean_name"])
                        if ball > eng_yuqori_ball:
                            eng_yuqori_ball = ball
                            eng_yaxshi_moslik = talaba

                    if eng_yaxshi_moslik and eng_yuqori_ball >= 75:
                        target_row = eng_yaxshi_moslik["row"]
                        eski_summa_val = sheet_read.cell(row=target_row, column=5).value
                        try:
                            eski_summa = float(eski_summa_val) if eski_summa_val else 0.0
                        except (ValueError, TypeError):
                            eski_summa = 0.0

                        jami_yangi = eski_summa + yangi_summa
                        sheet_write.cell(row=target_row, column=5).value = jami_yangi
                        
                        yangilanish_tarixi.append(
                            f"👤 **{eng_yaxshi_moslik['original_name']}**\n"
                            f"├ 🏦 Bankda: `{deb_fio}`\n"
                            f"├ ➕ Tushgan pul: `{yangi_summa:,.0f} so'm`\n"
                            f"└ 📊 Baza: `{eski_summa:,.0f}` ➔ Yangi Jami: `{jami_yangi:,.0f} so'm`"
                        )
                    else:
                        topilmaganlar.append(f"❓ `{deb_fio}` — {yangi_summa:,.0f} so'm")

            wb_baza_write.save(natija_nomi)
            wb_baza_write.close()
            wb_baza_read.close()
            wb_deb.close()

            hisobot_matni = f"📊 **KONTRAKT YANGILANISH HISOBOTI**\n"
            hisobot_matni += f"📅 Filtr sanasi: {cheklov_sanasi.strftime('%d.%m.%Y')} dan boshlab\n"
            hisobot_matni += f"━━━━━━━━━━━━━━━━━━━━\n\n"
            
            if yangilanish_tarixi:
                hisobot_matni += "✅ **Yangilangan talabalar ro'yxati:**\n\n"
                for t in yangilanish_tarixi[:12]:
                    hisobot_matni += t + "\n\n"
                if len(yangilanish_tarixi) > 12:
                    hisobot_matni += f"i... va yana {len(yangilanish_tarixi)-12} ta talaba yangilandi.\n\n"
            else:
                hisobot_matni += "❌ Yangi to'lovlar topilmadi.\n\n"

            if topilmaganlar:
                hisobot_matni += "⚠️ **Ismi topilmagan bank to'lovlari:**\n"
                for top in topilmaganlar[:8]:
                    hisobot_matni += top + "\n"

            with open(natija_nomi, 'rb') as f_send:
                bot.send_document(chat_id, f_send, caption="📄 Formulalari buzilmagan tayyor Excel faylingiz.")
            
            bot.send_message(chat_id, hisobot_matni)

            if os.path.exists(baza_path): os.remove(baza_path)
            if os.path.exists(deb_nomi): os.remove(deb_nomi)
            if os.path.exists(natija_nomi): os.remove(natija_nomi)
            user_data[chat_id] = {}
            print(f"✅ PROCESSING COMPLETE for {chat_id}")

    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        bot.send_message(chat_id, f"❌ Xatolik yuz berdi: {str(e)}")
        user_data[chat_id] = {}

@bot.message_handler(func=lambda message: True)
def handle_other_messages(message):
    """Boshqa xabarlar uchun default handler"""
    chat_id = message.chat.id
    print(f"📨 OTHER MESSAGE from {chat_id}: {message.text[:50] if message.text else 'no text'}")
    bot.send_message(chat_id, "❌ Noma'lum buyruq. /start bilan boshlang!")

# ============ WEBHOOK ENDPOINTS ============
@app.route('/' + TOKEN, methods=['POST'])
def getMessage():
    print(f"\n📨 WEBHOOK RECEIVED AT /{TOKEN}")
    try:
        json_data = request.get_json()
        print(f"📦 JSON DATA: {json.dumps(json_data, indent=2)[:200]}...")
        
        update = telebot.types.Update.de_json(json_data)
        print(f"✅ UPDATE OBJECT CREATED: {update}")
        
        # Directly handle message
        if update.message:
            print(f"✅ MESSAGE DETECTED: {update.message}")
            bot.process_new_updates([update])
            print(f"✅ MESSAGE PROCESSED")
        else:
            print(f"❌ NO MESSAGE IN UPDATE")
            
    except Exception as e:
        print(f"❌ WEBHOOK ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    return "ok", 200

@app.route("/")
def webhook():
    return "Bot Render bulutida 24/7 faol! ✅", 200

@app.route("/health")
def health():
    return "OK", 200

# ============ WEBHOOK SETUP ============
def setup_webhook():
    """Webhook o'rnatish"""
    try:
        if not TOKEN or not WEBHOOK_URL:
            print("❌ TOKEN yoki WEBHOOK_URL topilmadi!")
            return
        
        print(f"\n🔧 WEBHOOK SETUP STARTING...")
        bot.remove_webhook()
        print(f"✅ OLD WEBHOOK REMOVED")
        
        webhook_full_url = WEBHOOK_URL + "/" + TOKEN
        print(f"✅ WEBHOOK URL: {webhook_full_url}")
        
        bot.set_webhook(url=webhook_full_url)
        print(f"✅ WEBHOOK SET SUCCESSFULLY!")
        
        # Tekshirish
        info = bot.get_webhook_info()
        print(f"✅ WEBHOOK INFO: {info}")
        print(f"🚀 BOT READY TO RECEIVE MESSAGES!\n")
        
    except Exception as e:
        print(f"❌ WEBHOOK SETUP ERROR: {e}")
        import traceback
        traceback.print_exc()

# App startup
setup_webhook()

# ============ MAIN ============
if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    print(f"🚀 SERVER STARTING ON PORT {port}...")
    app.run(host="0.0.0.0", port=port, debug=False)
